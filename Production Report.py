import xlrd
location=('C:/Users/lrzhang/Desktop/Python sheets/S1141.xlsx')
wb = xlrd.open_workbook(location)
sheet = wb.sheet_by_index(0)
Station_number=sheet.name[1]+sheet.name[2]+sheet.name[3]+sheet.name[4]
Station_number=int(Station_number)

Station_row=0
Reject_per_shift=0
Total_reject=0
Good_per_shift=0
Total_good=0
Total_parts=0
OA=0
FTT=0
D_shift_start=(6.5/24) %1
A_shift_start=(14.5/24) %1
N_shift_start=(20.5/24) %1
D_run=0
A_run=0
N_run=0
Timestamp=0.0

class Cell:
    row=0
    column=0
class Part:
    Type_Name=''
    Total_Run=0
    Good=0
    Total_Reject=0
    Run_D=0
    Run_A=0
    Run_N=0
    Good_D=0
    Good_A=0
    Good_N=0
    Reject_D=0
    Reject_A = 0
    Reject_N = 0

Reject_code=Cell()
Count=Cell()
Percentage=Cell()
FTT=Cell()
STT=Cell()
Part1=Part()
Part2=Part()
Part3=Part()
Part4=Part()
Part5=Part()
#To find where the Reject code cell starts
for row in range(sheet.nrows):
    for column in range(sheet.ncols):
        if sheet.cell_value(row,column)=='Reject code':
            Reject_code.row=row+1
            Reject_code.column=column
#To find where the Count cell starts
Count.row= Reject_code.row
Count.column=Reject_code.column+1
Percentage.row= Reject_code.row
Percentage.column=Reject_code.column+2
FTT.row= Reject_code.row
FTT.column=Reject_code.column+4
STT.row= Reject_code.row
STT.column=Reject_code.column+5
#To find total reject
'''for row in range(bar_code.row, sheet.nrows):
    if sheet.cell_value(row,status_code.column)!=11410:
       Total_reject = Total_reject + 1
print(Total_reject)'''

#To find CSS run count, include good, reject and total, in different shift
for row in range(bar_code.row, sheet.nrows):
    #Total_parts = Total_parts + 1
    Timestamp = sheet.cell_value(row, 0)
    Timestamp = Timestamp % 1
    if sheet.cell_value(row,bar_code.column)[0]=='4':
      if (Timestamp>= D_shift_start) and (Timestamp< A_shift_start):
        Part1.Run_D=Part1.Run_D+1
        if sheet.cell_value(row,status_code.column)!=11410:
            Part1.Reject_D=Part1.Reject_D+1
      if (Timestamp>= A_shift_start) and (Timestamp< N_shift_start):
        Part1.Run_A=Part1.Run_A+1
        if sheet.cell_value(row,status_code.column)!=11410:
            Part1.Reject_A=Part1.Reject_A+1
      if ((Timestamp>= N_shift_start) and (Timestamp< 1.0)) or ((Timestamp>= 0.0) and (Timestamp< D_shift_start)):
        Part1.Run_N=Part1.Run_N+1
        if sheet.cell_value(row,status_code.column)!=11410:
            Part1.Reject_N=Part1.Reject_N+1
      Part1.Type_Name='CSS'
      Part1.Total_Run=Part1.Total_Run+1
      Part1.Good=Part1.Total_Run-Part1.Total_Reject
      if sheet.cell_value(row,status_code.column)!=11410:
        Part1.Total_Reject=Part1.Total_Reject+1

#To find 10R60
    if sheet.cell_value(row,bar_code.column)[0]=='1' and sheet.cell_value(row,bar_code.column)[15]=='B' :
      if (Timestamp>= D_shift_start) and (Timestamp< A_shift_start):
        Part2.Run_D=Part2.Run_D+1
        if sheet.cell_value(row,status_code.column)!=11410:
            Part2.Reject_D=Part2.Reject_D+1
      if (Timestamp>= A_shift_start) and (Timestamp< N_shift_start):
        Part2.Run_A=Part2.Run_A+1
        if sheet.cell_value(row,status_code.column)!=11410:
            Part2.Reject_A=Part2.Reject_A+1
      if ((Timestamp>= N_shift_start) and (Timestamp< 1.0)) or ((Timestamp>= 0.0) and (Timestamp< D_shift_start)):
        Part2.Run_N=Part2.Run_N+1
        if sheet.cell_value(row,status_code.column)!=11410:
            Part2.Reject_N=Part2.Reject_N+1
      Part2.Type_Name='10R60'
      Part2.Total_Run = Part2.Total_Run + 1
      Part2.Good = Part2.Total_Run - Part2.Total_Reject
      if sheet.cell_value(row,status_code.column)!=11410:
        Part2.Total_Reject=Part2.Total_Reject+1

#To find AB1V
    if sheet.cell_value(row,bar_code.column)[0]=='1' and sheet.cell_value(row,bar_code.column)[15]=='C' :
      if (Timestamp>= D_shift_start) and (Timestamp< A_shift_start):
        Part3.Run_D=Part3.Run_D+1
        if sheet.cell_value(row,status_code.column)!=11410:
            Part3.Reject_D=Part3.Reject_D+1
      if (Timestamp>= A_shift_start) and (Timestamp< N_shift_start):
        Part3.Run_A=Part3.Run_A+1
        if sheet.cell_value(row,status_code.column)!=11410:
            Part3.Reject_A=Part3.Reject_A+1
      if ((Timestamp>= N_shift_start) and (Timestamp< 1.0)) or ((Timestamp>= 0.0) and (Timestamp< D_shift_start)):
        Part3.Run_N=Part3.Run_N+1
        if sheet.cell_value(row,status_code.column)!=11410:
            Part3.Reject_N=Part3.Reject_N+1
      Part3.Type_Name='AB1V'
      Part3.Total_Run = Part3.Total_Run + 1
      Part3.Good = Part3.Total_Run - Part3.Total_Reject
      if sheet.cell_value(row,status_code.column)!=11410:
        Part3.Total_Reject=Part3.Total_Reject+1

#To find Gen2
    if sheet.cell_value(row,bar_code.column)[0]=='1' and sheet.cell_value(row,bar_code.column)[15]=='D' :
      if (Timestamp>= D_shift_start) and (Timestamp< A_shift_start):
        Part4.Run_D=Part4.Run_D+1
        if sheet.cell_value(row,status_code.column)!=11410:
            Part4.Reject_D=Part4.Reject_D+1
      if (Timestamp>= A_shift_start) and (Timestamp< N_shift_start):
        Part4.Run_A=Part4.Run_A+1
        if sheet.cell_value(row,status_code.column)!=11410:
            Part4.Reject_A=Part4.Reject_A+1
      if ((Timestamp>= N_shift_start) and (Timestamp< 1.0)) or ((Timestamp>= 0.0) and (Timestamp< D_shift_start)):
        Part4.Run_N=Part4.Run_N+1
        if sheet.cell_value(row,status_code.column)!=11410:
            Part4.Reject_N=Part4.Reject_N+1
      Part4.Type_Name='GEN2'
      Part4.Total_Run = Part4.Total_Run + 1
      Part4.Good = Part4.Total_Run - Part4.Total_Reject
      if sheet.cell_value(row,status_code.column)!=11410:
        Part4.Total_Reject=Part4.Total_Reject+1
'''print(Total_parts)
if Part1.Type_Name!='':
 print(Part1.Type_Name,Part1.Total_Reject,Part1.Total_Run,Part1.Run_D,Part1.Run_A,Part1.Run_N)
if Part2.Type_Name!='':
 print(Part2.Type_Name,Part2.Total_Reject,Part2.Total_Run,Part2.Run_D,Part2.Run_A,Part2.Run_N)
if Part3.Type_Name!='':
 print(Part3.Type_Name,Part3.Total_Reject,Part3.Total_Run,Part3.Run_D,Part3.Run_A,Part3.Run_N)
if Part4.Type_Name != '':
 print(Part4.Type_Name,Part4.Total_Reject,Part4.Total_Run,Part4.Run_D,Part4.Run_A,Part4.Run_N)

# Total run parts for different shift
for row in range (3,sheet.nrows):
    Timestamp=sheet.cell_value(row,0)
    Timestamp=Timestamp % 1
    if (Timestamp>= D_shift_start) and (Timestamp< A_shift_start):
        D_run=D_run+1
    if (Timestamp>= A_shift_start) and (Timestamp< N_shift_start):
        A_run=A_run+1
    if ((Timestamp>= N_shift_start) and (Timestamp< 1.0)) or ((Timestamp>= 0.0) and (Timestamp< D_shift_start)):
        N_run=N_run+1
#print(D_shift_start,A_shift_start,N_shift_start)
#print(sheet.cell_value(1588,0)%1)

print(D_run,A_run,N_run)'''
#################################################################################################
# Write to Excel Daily
# Initialization
from openpyxl import load_workbook

wb = load_workbook("C:/Users/lrzhang/Desktop/Python sheets/Daily.xlsx")
Month, Date = input('Report For Date:').split()
Sheet = wb.worksheets[int(Month) - 1]
New_date_column = 0
New_product_column = 1

# Find the row and column to record date and product
for row in range(1, Sheet.max_row):
    for column in range(1, Sheet.max_column):
        if Sheet.cell(row, column).value == 'DAY':
            DAY_row = row
for row in range(1, Sheet.max_row):
    for column in range(1, Sheet.max_column):
        if Sheet.cell(row, column).value == 'PRODUCT':
            Product_row = row
for column in range(2, Sheet.max_column):
    if Sheet.cell(23, column).value is None:
        New_date_column = column
        break
New_product_column = New_date_column
print(DAY_row,Product_row)
#To find station row for data writing
for row in range(1, Sheet.max_row):
    for column in range(1, Sheet.max_column):
        if Sheet.cell(row,column).value==1141:
            Station_row=row

# Write the new date input by user to cell
Sheet.cell(DAY_row, New_date_column).value = Date
# Write product run, divided into shift
if Part1.Run_D > 0:
    Sheet.cell(Product_row, New_product_column).value = Part1.Type_Name
    Sheet.cell(Product_row+1, New_product_column).value = 'D'
    Sheet.cell(Station_row, New_product_column).value = Part1.Reject_D
    New_product_column = New_product_column + 1

if Part2.Run_D > 0:
    Sheet.cell(Product_row, New_product_column).value = Part2.Type_Name
    Sheet.cell(Product_row+1, New_product_column).value = 'D'
    Sheet.cell(Station_row, New_product_column).value = Part2.Reject_D
    New_product_column = New_product_column + 1
if Part3.Run_D > 0:
    Sheet.cell(Product_row, New_product_column).value = Part3.Type_Name
    Sheet.cell(Product_row + 1, New_product_column).value = 'D'
    Sheet.cell(Station_row, New_product_column).value = Part3.Reject_D
    New_product_column = New_product_column + 1
if Part4.Run_D > 0:
    Sheet.cell(Product_row, New_product_column).value = Part4.Type_Name
    Sheet.cell(Product_row + 1, New_product_column).value = 'D'
    Sheet.cell(Station_row, New_product_column).value = Part4.Reject_D
    New_product_column = New_product_column + 1
if Part1.Run_A > 0:
    Sheet.cell(Product_row, New_product_column).value = Part1.Type_Name
    Sheet.cell(Product_row + 1, New_product_column).value = 'A'
    Sheet.cell(Station_row, New_product_column).value = Part1.Reject_A
    New_product_column = New_product_column + 1
if Part2.Run_A > 0:
    Sheet.cell(Product_row, New_product_column).value = Part2.Type_Name
    Sheet.cell(Product_row + 1, New_product_column).value = 'A'
    Sheet.cell(Station_row, New_product_column).value = Part2.Reject_A
    New_product_column = New_product_column + 1
if Part3.Run_A > 0:
    Sheet.cell(Product_row, New_product_column).value = Part3.Type_Name
    Sheet.cell(Product_row + 1, New_product_column).value = 'A'
    Sheet.cell(Station_row, New_product_column).value = Part3.Reject_A
    New_product_column = New_product_column + 1
if Part4.Run_A > 0:
    Sheet.cell(Product_row, New_product_column).value = Part4.Type_Name
    Sheet.cell(Product_row + 1, New_product_column).value = 'A'
    Sheet.cell(Station_row, New_product_column).value = Part4.Reject_A
    New_product_column = New_product_column + 1
if Part1.Run_N > 0:
    Sheet.cell(Product_row, New_product_column).value = Part1.Type_Name
    Sheet.cell(Product_row + 1, New_product_column).value = 'N'
    Sheet.cell(Station_row, New_product_column).value = Part1.Reject_N
    New_product_column = New_product_column + 1
if Part2.Run_N > 0:
    Sheet.cell(Product_row, New_product_column).value = Part2.Type_Name
    Sheet.cell(Product_row + 1, New_product_column).value = 'N'
    Sheet.cell(Station_row, New_product_column).value = Part2.Reject_N
    New_product_column = New_product_column + 1
if Part3.Run_N > 0:
    Sheet.cell(Product_row, New_product_column).value = Part3.Type_Name
    Sheet.cell(Product_row + 1, New_product_column).value = 'N'
    Sheet.cell(Station_row, New_product_column).value = Part3.Reject_N
    New_product_column = New_product_column + 1
if Part4.Run_N > 0:
    Sheet.cell(Product_row, New_product_column).value = Part4.Type_Name
    Sheet.cell(Product_row + 1, New_product_column).value = 'N'
    Sheet.cell(Station_row, New_product_column).value = Part4.Reject_N
    New_product_column = New_product_column + 1

for column in range (3,New_product_column):
    Reject_per_shift = 0
    for row in range (4,23):
     if Sheet.cell(row, column).value is not None:
       Reject_per_shift=Reject_per_shift+Sheet.cell(row, column).value
     Sheet.cell(23, column).value = Reject_per_shift
#print(New_product_column)
#print(Sheet.cell(Product_row, New_product_column).value)
wb.save("C:/Users/lrzhang/Desktop/Python sheets/Daily Trial.xlsx")

